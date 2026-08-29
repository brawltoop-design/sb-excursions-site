"""Две таблицы результата.

youtube-demand-map.xlsx — карта спроса: запрос × топ-20 × владелец слота.
ota-audit-data.xlsx — сторона OTA: реклама, свои каналы, партнёрские улики.
"""
import json,collections,statistics,datetime
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,PatternFill
from openpyxl.utils import get_column_letter

OUT="ota-audit/raw"
rows=json.load(open(f"{OUT}/yt-search-classified.json"))
ch=json.load(open(f"{OUT}/yt-channels.json"))
vids=json.load(open(f"{OUT}/yt-videos.json"))
ads=json.load(open(f"{OUT}/google-ads-transparency.json"))
own=json.load(open(f"{OUT}/ota-own-channels.json"))
ev=json.load(open(f"{OUT}/yt-affiliate-evidence.json"))
NOW=datetime.datetime(2026,8,29,tzinfo=datetime.timezone.utc)

HEAD=Font(bold=True,color="FFFFFF"); FILL=PatternFill("solid",fgColor="374151")
def style(ws,widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    for c in ws[1]:
        c.font=HEAD; c.fill=FILL; c.alignment=Alignment(vertical="center",wrap_text=True)
    ws.freeze_panes="A2"

RU={"OTA_OWN":"OTA, свой канал","CREATOR_OTA_AFFILIATE":"автор с партнёрской ссылкой OTA",
    "CREATOR_MENTIONS_OTA":"автор, упоминает OTA без метки","CREATOR_INDEPENDENT":"независимый автор",
    "LOCAL_OPERATOR":"местный оператор","LOCAL_OPERATOR_AFFILIATE":"местный оператор с партнёркой"}
LVL={"L1_inspiration":"1. вдохновение","L2_activity":"2. занятие","L3_booking":"3. бронирование"}

# ---------- карта спроса ----------
wb=Workbook(); ws=wb.active; ws.title="выдача"
ws.append(["уровень","запрос","позиция","канал","владелец","видео","просмотры",
           "подписчики","просм/подп","возраст, лет","страна канала","ссылка"])
for r in sorted(rows,key=lambda x:(x["level"],x["query"],x["position"])):
    c=ch.get(r["channel_id"],{}); v=vids.get(r["video_id"],{})
    subs=c.get("subscribers",0); vw=v.get("views",0)
    age=(NOW-datetime.datetime.fromisoformat(r["published_at"].replace("Z","+00:00"))).days/365.25
    ws.append([LVL.get(r["level"],r["level"]),r["query"],r["position"],c.get("title",""),
               RU.get(r["owner"],r["owner"]),r["title"][:90],vw,subs,
               round(vw/subs,2) if subs else None,round(age,1),c.get("country",""),
               f"https://youtu.be/{r['video_id']}"])
style(ws,[15,30,8,28,30,50,12,12,10,11,12,32])

ws2=wb.create_sheet("итоги по запросам")
ws2.append(["уровень","запрос","OTA свой канал","авторы с партнёркой OTA","местные операторы",
            "независимые авторы","сумма просмотров топ-20","медиана просмотров","медиана возраста, лет"])
byq=collections.defaultdict(list)
for r in rows: byq[(r["level"],r["query"])].append(r)
for (lvl,q),rs in sorted(byq.items()):
    cn=collections.Counter(r["owner"] for r in rs)
    vw=[vids[r["video_id"]]["views"] for r in rs if r["video_id"] in vids]
    ag=[(NOW-datetime.datetime.fromisoformat(r["published_at"].replace("Z","+00:00"))).days/365.25 for r in rs]
    ws2.append([LVL.get(lvl,lvl),q,cn["OTA_OWN"],
                cn["CREATOR_OTA_AFFILIATE"]+cn["LOCAL_OPERATOR_AFFILIATE"],
                cn["LOCAL_OPERATOR"]+cn["LOCAL_OPERATOR_AFFILIATE"],cn["CREATOR_INDEPENDENT"],
                sum(vw),round(statistics.median(vw)) if vw else 0,round(statistics.median(ag),1)])
style(ws2,[15,34,15,20,17,17,20,17,18])

ws3=wb.create_sheet("каналы")
ws3.append(["канал","владелец","слотов в топ-20","подписчики","всего видео","страна",
            "партнёрские ссылки OTA","партнёрские сети"])
app=collections.Counter(r["channel_id"] for r in rows)
cls={r["channel_id"]:r["owner"] for r in rows}
for cid,n in app.most_common():
    c=ch.get(cid,{}); e=ev.get(cid,{})
    ws3.append([c.get("title",""),RU.get(cls.get(cid),""),n,c.get("subscribers",0),
                c.get("video_count",0),c.get("country",""),
                ", ".join(e.get("ota",{}).keys()),", ".join(e.get("net",{}).keys())])
style(ws3,[34,30,15,13,12,9,26,20])
wb.save("ota-audit/youtube-demand-map.xlsx")

# ---------- сторона OTA ----------
wb2=Workbook(); a=wb2.active; a.title="реклама"
a.append(["бренд","регион","всего объявлений","из них на YouTube","доля YouTube","формат креативов"])
fmt=ads["формат_креативов"]
for m in ads["замеры"]:
    a.append([m["бренд"],m["регион"],m["всего"],m["youtube"],
              f'{m["доля_youtube"]*100:.0f}%' if m["доля_youtube"] else "нет данных",
              fmt.get(f'{m["бренд"]}_{m["регион"]}',fmt.get(m["бренд"],""))])
a.append([])
a.append(["источник",ads["источник"],"снято",ads["снято"]])
a.append(["оговорка",ads["оговорка"]])
style(a,[16,10,18,20,14,60])

b=wb2.create_sheet("свои каналы OTA")
b.append(["бренд","канал","подписчики","всего видео","всего просмотров",
          "просмотров на подписчика","дней с последней загрузки","медиана просмотров свежих 25"])
for k,d in sorted(own.items(),key=lambda x:-(x[1].get("total_views",0))):
    if "subs" not in d:
        b.append([k,"канал достоверно не определён","нет данных","нет данных","нет данных","нет данных","нет данных","нет данных"]); continue
    b.append([k,d["title"],d["subs"],d["video_count"],d["total_views"],
              round(d["total_views"]/max(d["subs"],1)),
              d.get("days_since_last_upload","нет данных"),d.get("median_views_last25","нет данных")])
style(b,[14,22,13,12,17,24,26,26])

c=wb2.create_sheet("партнёрские улики")
c.append(["канал","слотов в топ-20","подписчики","ссылки на OTA","партнёрские сети",
          "видео с партнёрской меткой","вывод"])
for cid,n in app.most_common():
    e=ev.get(cid,{})
    if not e.get("ota") and not e.get("net"): continue
    ci=ch.get(cid,{})
    verdict="подтверждённый партнёр OTA" if (e.get("net") or e.get("tags")) else "упоминает OTA, партнёрство не доказано"
    c.append([ci.get("title",""),n,ci.get("subscribers",0),", ".join(e.get("ota",{}).keys()),
              ", ".join(e.get("net",{}).keys()),e.get("tags",0),verdict])
style(c,[34,16,13,26,20,26,36])
wb2.save("ota-audit/ota-audit-data.xlsx")
print("готово: youtube-demand-map.xlsx, ota-audit-data.xlsx")
