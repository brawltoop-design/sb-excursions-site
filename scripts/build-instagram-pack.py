# -*- coding: utf-8 -*-
"""Сборка пакета для Instagram: таблица постов + папки с медиа.

На выходе:
  instagram-pack/posts.xlsx          — все карточки, тексты en/ru, пути к файлам
  instagram-pack/<id>/01-...jpg      — медиа поста, пронумерованы по порядку слайдов
  instagram-pack/README.md           — как этим пользоваться

Медиа копируем, а не ссылаемся: заказчик открывает папку и тащит файлы в Figma,
не разбираясь, где что лежит в репозитории.
"""
import json, io, os, re, shutil, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "instagram-pack")
posts = json.load(io.open(sys.argv[1], encoding="utf8"))
if isinstance(posts, dict):
    posts = posts.get("posts", [])

os.makedirs(OUT, exist_ok=True)
HEAD = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F2937")
TYPE_FILL = {
    "тур": PatternFill("solid", fgColor="FEF3C7"),
    "информационный": PatternFill("solid", fgColor="DBEAFE"),
    "развлекательный": PatternFill("solid", fgColor="FCE7F3"),
}

wb = Workbook()
ws = wb.active
ws.title = "карточки"
ws.append(["пост", "тип", "название", "слайд", "роль", "заголовок EN", "текст EN",
           "заголовок RU", "текст RU", "файл медиа", "видео", "подсказка для Figma"])

copied = missing = 0
for p in posts:
    pid = re.sub(r"[^a-z0-9-]", "", str(p.get("id", "post")).lower()) or "post"
    pdir = os.path.join(OUT, pid)
    os.makedirs(pdir, exist_ok=True)
    for s in sorted(p.get("slides", []), key=lambda x: x.get("n", 0)):
        src_rel = str(s.get("image", "") or "")
        src = os.path.join(REPO, src_rel)
        name = ""
        if src_rel and os.path.isfile(src):
            ext = os.path.splitext(src_rel)[1]
            name = f"{s.get('n', 0):02d}-{os.path.splitext(os.path.basename(src_rel))[0]}{ext}"
            dst = os.path.join(pdir, name)
            if not os.path.exists(dst):
                shutil.copy2(src, dst)
            copied += 1
        elif src_rel:
            missing += 1
            name = f"НЕТ ФАЙЛА: {src_rel}"
        r = ws.max_row + 1
        ws.append([pid, p.get("type", ""), p.get("title_ru", ""), s.get("n", 0), s.get("role", ""),
                   s.get("headline_en", ""), s.get("body_en", ""),
                   s.get("headline_ru", ""), s.get("body_ru", ""),
                   name, "да" if s.get("is_video") else "", s.get("design_note", "")])
        f = TYPE_FILL.get(p.get("type"))
        if f:
            ws.cell(row=r, column=2).fill = f

for i, w in enumerate([16, 16, 30, 7, 12, 30, 46, 30, 46, 30, 6, 34], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for c in ws[1]:
    c.font = HEAD; c.fill = FILL; c.alignment = Alignment(vertical="center", wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"

w2 = wb.create_sheet("посты")
w2.append(["пост", "тип", "название", "цель", "тур", "слайдов", "подпись EN", "подпись RU",
           "хэштеги", "идея для рилса"])
for p in posts:
    w2.append([p.get("id", ""), p.get("type", ""), p.get("title_ru", ""), p.get("goal", ""),
               p.get("tour_slug", ""), len(p.get("slides", [])),
               p.get("caption_en", ""), p.get("caption_ru", ""),
               p.get("hashtags", ""), p.get("reel_idea", "")])
for i, w in enumerate([16, 16, 30, 40, 24, 8, 50, 50, 34, 54], 1):
    w2.column_dimensions[get_column_letter(i)].width = w
for c in w2[1]:
    c.font = HEAD; c.fill = FILL; c.alignment = Alignment(vertical="center", wrap_text=True)
for row in w2.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
w2.freeze_panes = "A2"

wb.save(os.path.join(OUT, "posts.xlsx"))
print(json.dumps({"постов": len(posts),
                  "карточек": ws.max_row - 1,
                  "файлов скопировано": copied,
                  "файлов не найдено": missing}, ensure_ascii=False, indent=1))
